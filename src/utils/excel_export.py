import pandas as pd
from typing import Dict, List
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def export_summary_to_excel(
    df: pd.DataFrame,
    excel_path: str,
    column_mapping: Dict[str, str] = {
        "target_company": "Target Company",
        "group": "Group",
        "group_hq": "Group HQ",
        "vertical": "Vertical",
        "potential_disposal": "Potential disposal (s)",
        "potential_disposal_company": "Disposal company",
        "potential_disposal_country": "Disposal country",
        "disposal_nc_sector": "Disposal NC Sector",
        "rationale": "Rationale",
        "date": "Date",
        "interest_score": "Score",
        "carve_out_stage": "Carve Out Stage",
        "article_quote": "Article quote",
        "title": "Article Title",
        "source_name": "Source Name",
        "article_fragment": "Article Fragment",
    },
    sort_by: List[str] = ["Score", "Date"],
) -> str:
    df_export = df.rename(columns=column_mapping)
    df_export = df_export.sort_values(sort_by, ascending=False)

    # Export to Excel without index column
    df_export.to_excel(excel_path, index=False, sheet_name="Carveouts Summary")

    wb = load_workbook(excel_path)
    ws = wb['Carveouts Summary']  

    # Set Arial font for entire worksheet - slightly larger font (10.5pt)
    arial_font = Font(name="Arial", size=11)
    for row in ws.rows: #type: ignore
        for cell in row:
            cell.font = arial_font

    # Style headers - almost black background with white text
    header_fill = PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=12)

    for cell in ws[1]: #type: ignore
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Adjust column widths based on the screenshot
    ws.column_dimensions[get_column_letter(1)].width = 28  # Target Company 
    ws.column_dimensions[get_column_letter(2)].width = 18  # Group
    ws.column_dimensions[get_column_letter(3)].width = 12  # Group HQ (narrow)
    ws.column_dimensions[get_column_letter(4)].width = 16  # Vertical
    ws.column_dimensions[get_column_letter(5)].width = 30  # Potential disposal
    ws.column_dimensions[get_column_letter(6)].width = 28  # Disposal Company Name
    ws.column_dimensions[get_column_letter(7)].width = 20  # Disposal Country
    ws.column_dimensions[get_column_letter(8)].width = 28  # Disposal NC Sector
    ws.column_dimensions[get_column_letter(9)].width = 45  # Rationale (wider)
    ws.column_dimensions[get_column_letter(10)].width = 12  # Date
    ws.column_dimensions[get_column_letter(11)].width = 12  # Interest Score (narrow)
    ws.column_dimensions[get_column_letter(12)].width = 16  # Carve Out Stage (narrow)
    ws.column_dimensions[get_column_letter(13)].width = 45  # Article Quote (wider)
    ws.column_dimensions[get_column_letter(14)].width = 20  # Article Title (wider)
    ws.column_dimensions[get_column_letter(15)].width = 30  # Article Source (wider)
    ws.column_dimensions[get_column_letter(16)].width = 65  # Article Fragment (very wide)

    # Add subtle alternating row colors - very light gray
    light_gray_fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:  # Even rows
            for cell in ws[row_idx]:
                cell.fill = light_gray_fill

    # Set text wrapping and vertical alignment for all data cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Format date column
    for row in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=row, column=10)  # Date column
        if date_cell.value:
            date_cell.number_format = "yyyy-mm-dd"
            date_cell.alignment = Alignment(horizontal="center", vertical="center")

        score_cell = ws.cell(row=row, column=11)  # Interest Score column
        if isinstance(score_cell.value, (int, float)):
            score_cell.number_format = "0.00"
            score_cell.alignment = Alignment(horizontal="center", vertical="center")

        co_stage_cell = ws.cell(row=row, column=9)
        if co_stage_cell.value:
            co_stage_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add very thin light gray borders to all cells
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Set row height for header
    ws.row_dimensions[1].height = 24

    # Set appropriate row heights for data rows based on content
    for row in range(2, ws.max_row + 1):
        # Make rows taller to accommodate wrapped text
        ws.row_dimensions[row].height = 80  # Increased height for better quote display

    wb.save(excel_path)
    return excel_path


# example usage:
# export_file = export_summary_to_excel(grouped_summary, f"../../extractions/{job_id}/carveouts_summary_grouped.xlsx")
